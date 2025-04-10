from datetime import datetime, timedelta, timezone
import os
from tkinter.font import Font

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook

import pytz

class writer():
    '''
    Class that handles writing and creating excel reports. 
    Contains only write_report()
    '''
    def __convert_to_datetime(self, string) -> datetime:
        try:
            # Parse string, handling timezone if present
            dt_obj = datetime.strptime(string, "%Y-%m-%d %H:%M:%S")
            return dt_obj
        except: #? Ok now bear with me, it may look stupid, be stupid and make me look stupid, but it works :) Did i mention how much i hate dates
            try:
                dt_obj = datetime.strptime(string, "%Y-%m-%d %H:%M")
                return dt_obj
            except ValueError as e:
                raise ValueError(f"Invalid datetime format: {string}") from e

    def __get_sk_date(self) -> str:
        bratislava_tz = pytz.timezone('Europe/Bratislava')
        # Ensure the datetime is in UTC before converting
        dt_obj = datetime.now()
        utc_time = dt_obj.replace(tzinfo=pytz.utc) if dt_obj.tzinfo is None else dt_obj.astimezone(pytz.utc)
        bratislava_time = utc_time.astimezone(bratislava_tz)  # Convert to Bratislava timezone
        return bratislava_time.strftime("%Y-%m-%d %H:%M:%S") 

    def __get_sk_date_str(self) -> str:
        # Ensure the datetime is in UTC before converting
        bratislava_tz = pytz.timezone('Europe/Bratislava')
        dt_obj = datetime.now()
        utc_time = dt_obj.replace(tzinfo=pytz.utc) if dt_obj.tzinfo is None else dt_obj.astimezone(pytz.utc)
        bratislava_time = utc_time.astimezone(bratislava_tz)  # Convert to Bratislava timezone
        return bratislava_time.strftime("%Y-%m-%d %H:%M:%S") 

    def __compare_timeof(self, a_timeof, today) -> bool:
        timeof = self.__convert_to_datetime(string=a_timeof)
        diff = today - timeof
        # If the lease from date is a minute behind the current date, dont allow the lease
        # This gives the user 2 minutes to make a reservation, before being time blocked by leasing into the past
        if (diff.total_seconds()/60) >= 2:
            return True

    def __get_latest_file(self, folder_path, use_modification_time=True):
        try:
            if not os.path.exists(folder_path):
                raise FileNotFoundError(f"The folder '{folder_path}' does not exist.")
            
            if not os.path.isdir(folder_path):
                raise NotADirectoryError(f"'{folder_path}' is not a directory.")

            latest_file = None
            latest_time = 0

            # Use os.scandir for better performance
            with os.scandir(folder_path) as entries:
                for entry in entries:
                    if entry.is_file():
                        # Use modification time or creation time based on the parameter
                        file_time = entry.stat().st_mtime 
                        
                        if file_time > latest_time:
                            latest_time = file_time
                            latest_file = entry.path

            return latest_file

        except (FileNotFoundError, NotADirectoryError, PermissionError, OSError) as e:
            print(f"An error occurred: {e}")
            return None


    def write_report(self, recipient:str , car_name:str, stk:str, drive_type:str, timeof:str, timeto:str):
        """
        Tries to write to an existing excel file located at os.getcwd()/reports
        If no file exists it will create it and name it in the format: {CURRENT_DATE}_NW_ICLS_report.xlsx else {CURRENT_DATE}_EXCEL_ICLS_report.xlsx \n
        
        If and error occurs during creation it will also produce an error log for the current day in the directory /reports .  \n
        
        The function itself checks if a new month/day has begun.\n
        If a new month began a new excel file is made and written to. \n
        If a new day began a new worksheet is made witht he days value as its name in the latest excel file in the directory /reports.

        Arguments:
        recipient:  str, who has leased the car 
        car_name:   str, which car had been leased 
        stk:        str, cars STK 
        drive_type: str, concataned string of Gas, Shifter_type (Diesel, Automatic) 
        timeof:     str, date format: %Y-%m-%d %H:%M:%S or %Y-%m-%d %H:%M 
        timeto:     str, date format: %Y-%m-%d %H:%M:%S or %Y-%m-%d %H:%M 

        
        """

                # Define styles
        red_flag_ft = Font(bold=True, color="B22222")
        red_flag_fill = PatternFill("solid", "B22222")
        Header_fill = PatternFill("solid", "00CCFFFF")
        Header_ft = Font(bold=True, color="000000", size=20)
        Data_ft = Font(size=17)  # New font for data cells

        Header_border = Border(
            left=Side(border_style="medium", color='FF000000'),
            right=Side(border_style="medium", color='FF000000'),
            top=Side(border_style="medium", color='FF000000'),
            bottom=Side(border_style="medium", color='FF000000')
        )

        header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        # To fix the wierd seconds missing error, i will just get rid of the seconds manually
        if timeof.count(":") > 1:
            timeof = timeof[:-3]
            
        if timeto.count(":") > 1:
            timeto = timeto[:-3]

        
        latest_file = self.__get_latest_file(f"{os.getcwd()}/reports")

        # Use year and month to check if a new excel spreadsheet needs to be created
        # '2025-01-21 15:37:00ICLS_report.csv'  '2025-01-21 15:37:26_ICLS_report.csv'
        try:
            #! New: 2025-02-27 09:13:38_NW_ICLS_report.xlsx -> 2025-01-21 18:53:46

            split_date = latest_file.split("_")
            str_date = split_date[0].replace("/app/reports/", "")  #!  2025-02-27 09:13:38

            # Invalid datetime format: /app/reports/2025-03-01 20:49:40

            file_dt_date = self.__convert_to_datetime(str_date)
            file_year = file_dt_date.year
            file_month = file_dt_date.month


            current_date = self.__convert_to_datetime(self.__get_sk_date())
            cur_year = current_date.year
            cur_month = current_date.month
            
            
            
            #timeof = timeof.strftime("%Y-%m-%d %H:%M:%S")
            if file_year == cur_year and file_month == cur_month:
                wb = openpyxl.load_workbook(latest_file)
                # If a sheet name has been made before compare it with today, if its not equal create a new worksheet with the new days number
                all_sheets = wb.sheetnames
                
                tm = self.__get_sk_date()
                # This should be a day but it does not work for some reason, just returns an integer
                cur_date = self.__convert_to_datetime(tm)

                if int(all_sheets[-1]) == cur_date.day: #! The same day 
                    # Select the last sheet, that should correspond to the current day
                    ws = wb[wb.sheetnames[-1]]
                    data = [["","",timeof, timeto, car_name, stk, drive_type, recipient, "NULL", "NULL", "NULL","NULL","NULL","NULL","NULL","NULL"]]
                    for row in data:
                        ws.append(row)

                    wb.save(latest_file)
                else: #! A new day had begun
                    # You need to recreate the header and data formating, as a new worksheet would be empty
                    # Also add the filler column values and stuff, save to the existing file
                    ws = wb.create_sheet(f"{cur_date.day}")

                    filler = ["","","","","","","",""]
                    data = [filler,filler,["", "", "Čas od", "Čas do", "Auto", "SPZ", "Typ","Email", "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie","Zašpinené", "Pš.Interiér", "Pš.Exterier", "Kolizia"],["","",timeof, timeto, car_name, stk, drive_type, recipient, "NULL","NULL","NULL","NULL","NULL","NULL","NULL","NULL"]]

                    for row in data:
                        ws.append(row)

                    red_flag_cell = ws["B3"]
                    red_flag_cell.font = red_flag_ft
                    red_flag_cell.fill = red_flag_fill
                    red_flag_cell.border = Header_border
                    
                    # Set row height for header row
                    ws.row_dimensions[3].height = 35
                    # Set column widths for data columns
                    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
                        ws.column_dimensions[col].width = 23
                    # Format header row (C3:J3)
                    for row_cells in ws["C3:K3"]:
                        for cell in row_cells:
                            cell.font = Header_ft
                            cell.alignment = header_alignment
                            cell.fill = Header_fill
                            cell.border = Header_border


                    wb.save(latest_file)

            else:
                wb = Workbook()
                del wb["Sheet"]
                ws = wb.create_sheet(f"{self.__convert_to_datetime(self.__get_sk_date_str()).day}")
                #email_ft = Font(bold=True, color="B22222")
                filler = ["","","","","","","",""]
                data = [filler,filler,["", "", "Čas od", "Čas do", "Auto", "SPZ", "Typ","Email", "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie","Zašpinené", "Pš.Interiér", "Pš.Exterier", "Kolizia"],["","",timeof, timeto, car_name, stk, drive_type, recipient, "NULL","NULL","NULL"]]

                for row in data:
                    ws.append(row)
                    # Format red flag cell (B3)
                red_flag_cell = ws["B3"]
                red_flag_cell.font = red_flag_ft
                red_flag_cell.fill = red_flag_fill
                red_flag_cell.border = Header_border
                email_cell = ws["B3"]
                #email_cell.font = email_ft

                # Set row height for header row
                ws.row_dimensions[3].height = 35

                # Set column widths for data columns
                for col in ["C", "D", "E", "F", "G", "H", "I", "J","K", "L", "M", "N", "O", "P"]:
                    ws.column_dimensions[col].width = 23

                # Format header row (C3:J3)
                for row_cells in ws["C3:K3"]:
                    for cell in row_cells:
                        cell.font = Header_ft
                        cell.alignment = header_alignment
                        cell.fill = Header_fill
                        cell.border = Header_border

                # Format data rows (from row 4 onwards, columns C-J)
                # for row in ws.iter_rows(min_row=4, min_col=3, max_col=10):
                #     for cell in row:
                #         cell.font = Data_ft
                # Set row height for data rows (from row 4 to the last row)
                wb.save(f"{os.getcwd()}/reports/{self.__get_sk_date()}_EXCEL_ICLS_report.xlsx")

        except Exception as e: #? ONLY HAPPENDS IF THE DIRECTORY IS EMPTY, SO LIKE ONCE
            with open(f"{os.getcwd()}/reports/{self.__get_sk_date()}_ERRORt.txt", "a+") as file:
                file.write(f"{e}")

            # Define styles
            red_flag_ft = Font(bold=True, color="B22222")
            red_flag_fill = PatternFill("solid", "B22222")
            Header_fill = PatternFill("solid", "00CCFFFF")
            Header_ft = Font(bold=True, color="000000", size=20)
            Data_ft = Font(size=17)  # New font for data cells
            Header_border = Border(left=Side(border_style="medium", color='FF000000'),right=Side(border_style="medium", color='FF000000'),top=Side(border_style="medium", color='FF000000'),bottom=Side(border_style="medium", color='FF000000'))
            header_alignment = Alignment(horizontal='center',vertical='center')

            wb = Workbook()
            del wb["Sheet"]
            ws = wb.create_sheet(f"{self.__convert_to_datetime(self.__get_sk_date_str()).day}")
            filler = ["","","","","","","",""]
            data = [filler,filler,["", "", "Čas od", "Čas do", "Auto", "SPZ", "TYP","Email", "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie","Zašpinené", "Pš.Interiér", "Pš.Exterier", "Kolizia"],["","",timeof, timeto, car_name, stk, drive_type, recipient,"NULL","NULL","NULL","NULL","NULL","NULL","NULL","NULL"]]
            for row in data:
                ws.append(row)
                # Format red flag cell (B3)
            red_flag_cell = ws["B3"]
            red_flag_cell.font = red_flag_ft
            red_flag_cell.fill = red_flag_fill
            red_flag_cell.border = Header_border
            email_cell = ws["B3"]
            # Set row height for header row
            ws.row_dimensions[3].height = 35
            # Set column widths for data columns
            for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
                ws.column_dimensions[col].width = 23
            # Format header row (C3:J3)
            for row_cells in ws["C3:K3"]:
                for cell in row_cells:
                    cell.font = Header_ft
                    cell.alignment = header_alignment
                    cell.fill = Header_fill
                    cell.border = Header_border
            # Format data rows (from row 4 onwards, columns C-J)
            # for row in ws.iter_rows(min_row=4, min_col=3, max_col=10):
            #     for cell in row:
            #         cell.font = Data_ft
                    # Set row height for data rows (from row 4 to the last row)
            wb.save(f"{os.getcwd()}/reports/{self.__get_sk_date()}_NW_ICLS_report.xlsx")