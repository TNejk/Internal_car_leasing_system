import os
import psycopg2
import pytz
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

class MonthlyExcelWriter:
    """
    Monthly batch Excel writer for car lease reports.
    Designed to run once per month automatically in Docker container.
    """
    
    def __init__(self):
        # Database configuration from environment variables
        self.db_host = os.getenv('DB_HOST')
        self.db_port = os.getenv('DB_PORT')
        self.db_user = os.getenv('POSTGRES_USER')
        self.db_pass = os.getenv('POSTGRES_PASS')
        self.db_name = os.getenv('POSTGRES_DB')
        
        # Slovakia timezone configuration
        self.bratislava_tz = pytz.timezone('Europe/Bratislava')
        
        # Excel styling to match original format
        self._init_styles()
        
    def _init_styles(self):
        """Initialize Excel styling constants for professional appearance."""
        # Original styles
        self.red_flag_ft = Font(bold=True, color="B22222")
        self.red_flag_fill = PatternFill("solid", "B22222")
        self.header_fill = PatternFill("solid", "00CCFFFF")
        self.header_ft = Font(bold=True, color="000000", size=20)
        self.data_ft = Font(size=17)
        
        # Enhanced professional styles
        self.title_font = Font(name='Calibri', size=28, bold=True, color="1F4E79")
        self.subtitle_font = Font(name='Calibri', size=16, bold=True, color="2F5597")
        self.stats_label_font = Font(name='Calibri', size=14, bold=True, color="404040")
        self.stats_value_font = Font(name='Calibri', size=14, color="1F4E79")
        self.section_header_font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        
        # Enhanced fills
        self.title_fill = PatternFill("solid", "E7F3FF")
        self.section_fill = PatternFill("solid", "4472C4")
        self.stats_fill = PatternFill("solid", "F2F8FF")
        self.alt_row_fill = PatternFill("solid", "F8F9FA")
        self.success_fill = PatternFill("solid", "D4EDDA")
        self.warning_fill = PatternFill("solid", "FFF3CD")
        self.danger_fill = PatternFill("solid", "F8D7DA")
        
        # Enhanced borders
        self.thick_border = Border(
            left=Side(border_style="thick", color='4472C4'),
            right=Side(border_style="thick", color='4472C4'),
            top=Side(border_style="thick", color='4472C4'),
            bottom=Side(border_style="thick", color='4472C4')
        )
        
        self.medium_border = Border(
            left=Side(border_style="medium", color='6C757D'),
            right=Side(border_style="medium", color='6C757D'),
            top=Side(border_style="medium", color='6C757D'),
            bottom=Side(border_style="medium", color='6C757D')
        )
        
        self.header_border = Border(
            left=Side(border_style="medium", color='FF000000'),
            right=Side(border_style="medium", color='FF000000'),
            top=Side(border_style="medium", color='FF000000'),
            bottom=Side(border_style="medium", color='FF000000')
        )
        
        # Enhanced alignment
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
    
    def connect_to_db(self) -> tuple:
        """Establish database connection."""
        try:
            db_con = psycopg2.connect(
                dbname=self.db_name, 
                user=self.db_user, 
                host=self.db_host, 
                port=self.db_port, 
                password=self.db_pass
            )
            cur = db_con.cursor()
            return db_con, cur
        except psycopg2.Error as e:
            print(f"ERROR: Database connection failed: {e}")
            return None, None
    
    def get_sk_date(self) -> datetime:
        """Get current time in Slovakia (Bratislava) timezone."""
        dt_obj = datetime.now()
        utc_time = dt_obj.replace(tzinfo=pytz.utc) if dt_obj.tzinfo is None else dt_obj.astimezone(pytz.utc)
        bratislava_time = utc_time.astimezone(self.bratislava_tz)
        return bratislava_time
    
    def convert_to_bratislava_timezone(self, dt_obj) -> str:
        """Convert datetime object to Slovakia timezone string in original format."""
        if dt_obj is None:
            return "NULL"
        
        # Handle timezone-aware datetime
        if dt_obj.tzinfo is None:
            utc_time = pytz.utc.localize(dt_obj)
        else:
            utc_time = dt_obj.astimezone(pytz.utc)
        
        bratislava_time = utc_time.astimezone(self.bratislava_tz)
        # Match original format: "25-02-2025 21:04"
        return bratislava_time.strftime("%d-%m-%Y %H:%M")
    
    def get_month_date_range(self, year: int, month: int) -> tuple:
        """Get start and end datetime for a specific month in Slovakia timezone."""
        # Start of month in Slovakia timezone
        start_date = datetime(year, month, 1, 0, 0, 0)
        
        # Start of next month
        if month == 12:
            end_date = datetime(year + 1, 1, 1, 0, 0, 0)
        else:
            end_date = datetime(year, month + 1, 1, 0, 0, 0)
        
        # Convert to timezone-aware datetimes in Bratislava timezone
        start_date = self.bratislava_tz.localize(start_date)
        end_date = self.bratislava_tz.localize(end_date)
        
        return start_date, end_date
    
    def get_lease_data_for_month(self, year: int, month: int) -> List[Dict[str, Any]]:
        """
        Get all lease data for a specific month.
        Includes leases that started in that month, regardless of when they ended.
        """
        conn, cur = self.connect_to_db()
        if not conn:
            raise RuntimeError("Could not establish database connection")
        
        try:
            start_date, end_date = self.get_month_date_range(year, month)
            
            query = """
                SELECT 
                    l.id_lease,
                    l.start_of_lease, 
                    l.end_of_lease, 
                    l.time_of_return, 
                    l.note, 
                    l.status,
                    l.private, 
                    l.car_health_check, 
                    l.dirty, 
                    l.exterior_damage, 
                    l.interior_damage, 
                    l.collision,
                    c.name AS car_name, 
                    c.stk, 
                    c.drive_type, 
                    c.gas,
                    d.email,
                    d.name AS driver_name       
                FROM lease AS l
                JOIN car AS c ON l.id_car = c.id_car
                JOIN driver AS d ON l.id_driver = d.id_driver
                WHERE l.start_of_lease >= %s AND l.start_of_lease < %s
                ORDER BY l.start_of_lease ASC;
            """
            
            cur.execute(query, (start_date, end_date))
            raw_data = cur.fetchall()
            
            # Convert to list of dictionaries for easier handling
            lease_data = []
            for row in raw_data:
                lease_record = {
                    'id_lease': row[0],
                    'start_of_lease': row[1],
                    'end_of_lease': row[2],
                    'time_of_return': row[3],
                    'note': row[4] or '',
                    'status': row[5],  # True = active/completed, False = cancelled
                    'private': row[6],
                    'car_health_check': row[7],
                    'dirty': row[8],
                    'exterior_damage': row[9],
                    'interior_damage': row[10],
                    'collision': row[11],
                    'car_name': row[12],
                    'stk': row[13],
                    'drive_type': row[14] or '',
                    'gas': row[15] or '',
                    'email': row[16],
                    'driver_name': row[17],
                    # Add calculated fields
                    'is_cancelled': not row[5],  # status = False means cancelled
                    'is_multi_month': self._is_multi_month_lease(row[1], row[2]),
                    'late_return': self._calculate_late_return(row[2], row[3]) if row[3] else False
                }
                lease_data.append(lease_record)
            
            return lease_data
            
        except Exception as e:
            print(f"ERROR: Failed to get lease data: {e}")
            raise
        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()
    
    def _is_multi_month_lease(self, start_date, end_date) -> bool:
        """Check if a lease spans multiple months."""
        if not start_date or not end_date:
            return False
        return start_date.month != end_date.month or start_date.year != end_date.year
    
    def _calculate_late_return(self, end_date, return_date) -> bool:
        """Calculate if a lease was returned late."""
        if not return_date or not end_date:
            return False
        return return_date > end_date 
    
    def create_excel_report(self, year: int, month: int, force_rebuild: bool = False) -> str:
        """
        Create monthly Excel report matching original format exactly.
        """
        # Create reports directory if it doesn't exist
        reports_dir = os.path.join(os.getcwd(), 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Generate filename: "2025.04 ICLS Report.xlsx"
        filename = f"{year}.{month:02d} ICLS Report.xlsx"
        filepath = os.path.join(reports_dir, filename)
        
        # Check if file already exists and force_rebuild is False
        if os.path.exists(filepath) and not force_rebuild:
            print(f"INFO: Report already exists: {filename}")
            return filename
        
        try:
            # Get lease data for the month
            lease_data = self.get_lease_data_for_month(year, month)
            print(f"INFO: Found {len(lease_data)} leases for {year}-{month:02d}")
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create summary sheet first
            self._create_summary_sheet(wb, lease_data, year, month)
            
            # Create leases data sheet (matching original format)
            self._create_leases_sheet(wb, lease_data, year, month)
            
            # Save workbook
            wb.save(filepath)
            print(f"INFO: Report created successfully: {filename}")
            
            return filename
            
        except Exception as e:
            print(f"ERROR: Failed to create Excel report: {e}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, lease_data: List[Dict], year: int, month: int):
        """Create enhanced summary statistics sheet with professional charts."""
        ws = wb.create_sheet("Prehľad", 0)
        
        # Set column widths for better layout
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # Main title with enhanced styling
        ws.merge_cells('B2:G2')
        title_cell = ws['B2']
        title_cell.value = f"📊 Mesačný prehľad leasing systému - {year}.{month:02d}"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment
        title_cell.fill = self.title_fill
        title_cell.border = self.thick_border
        ws.row_dimensions[2].height = 40
        
        # Calculate statistics
        total_leases = len(lease_data)
        completed_leases = len([l for l in lease_data if not l['is_cancelled'] and l['time_of_return']])
        cancelled_leases = len([l for l in lease_data if l['is_cancelled']])
        active_leases = len([l for l in lease_data if not l['is_cancelled'] and not l['time_of_return']])
        private_leases = len([l for l in lease_data if l['private']])
        business_leases = total_leases - private_leases
        multi_month_leases = len([l for l in lease_data if l['is_multi_month']])
        late_returns = len([l for l in lease_data if l['late_return']])
        
        # Damage statistics
        damaged_cars = len([l for l in lease_data if l['car_health_check']])
        dirty_cars = len([l for l in lease_data if l['dirty']])
        exterior_damage = len([l for l in lease_data if l['exterior_damage']])
        interior_damage = len([l for l in lease_data if l['interior_damage']])
        collisions = len([l for l in lease_data if l['collision']])
        
        # Section 1: Overall Statistics
        row = 4
        ws.merge_cells(f'B{row}:C{row}')
        section_cell = ws[f'B{row}']
        section_cell.value = "📋 Celkové štatistiky"
        section_cell.font = self.section_header_font
        section_cell.fill = self.section_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.medium_border
        ws.row_dimensions[row].height = 30
        
        # Overall stats with enhanced styling
        stats = [
            ("Celkový počet rezervácií:", total_leases, self.stats_fill),
            ("✅ Dokončené rezervácie:", completed_leases, self.success_fill),
            ("❌ Zrušené rezervácie:", cancelled_leases, self.danger_fill),
            ("🔄 Aktívne rezervácie:", active_leases, self.warning_fill),
            ("👤 Súkromné rezervácie:", private_leases, self.stats_fill),
            ("🏢 Firemné rezervácie:", business_leases, self.stats_fill),
            ("📅 Viacmesačné rezervácie:", multi_month_leases, self.stats_fill),
            ("⏰ Neskoré vrátenia:", late_returns, self.danger_fill if late_returns > 0 else self.success_fill),
        ]
        
        row += 1
        for label, value, fill in stats:
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = self.stats_label_font
            ws[f"B{row}"].fill = fill
            ws[f"B{row}"].border = self.medium_border
            ws[f"B{row}"].alignment = self.left_alignment
            
            ws[f"C{row}"] = value
            ws[f"C{row}"].font = self.stats_value_font
            ws[f"C{row}"].fill = fill
            ws[f"C{row}"].border = self.medium_border
            ws[f"C{row}"].alignment = self.center_alignment
            row += 1
        
        # Section 2: Damage Statistics
        row += 1
        ws.merge_cells(f'B{row}:C{row}')
        section_cell = ws[f'B{row}']
        section_cell.value = "🔧 Stav vozidiel"
        section_cell.font = self.section_header_font
        section_cell.fill = self.section_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.medium_border
        ws.row_dimensions[row].height = 30
        
        damage_stats = [
            ("🔍 Vyžaduje kontrolu:", damaged_cars),
            ("🧽 Znečistené vozidlá:", dirty_cars),
            ("🚪 Poškodený interiér:", interior_damage),
            ("🚗 Poškodený exteriér:", exterior_damage),
            ("💥 Kolízie:", collisions),
        ]
        
        row += 1
        for label, value in damage_stats:
            fill = self.danger_fill if value > 0 else self.success_fill
            
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = self.stats_label_font
            ws[f"B{row}"].fill = fill
            ws[f"B{row}"].border = self.medium_border
            ws[f"B{row}"].alignment = self.left_alignment
            
            ws[f"C{row}"] = value
            ws[f"C{row}"].font = self.stats_value_font
            ws[f"C{row}"].fill = fill
            ws[f"C{row}"].border = self.medium_border
            ws[f"C{row}"].alignment = self.center_alignment
            row += 1
        
        # Add charts
        self._add_lease_status_chart(ws, completed_leases, cancelled_leases, active_leases)
        self._add_lease_type_chart(ws, private_leases, business_leases)
        self._add_damage_chart(ws, damaged_cars, dirty_cars, exterior_damage, interior_damage, collisions)
        
        # Add summary insights
        self._add_summary_insights(ws, lease_data, row + 2)
    
    def _add_lease_status_chart(self, ws, completed: int, cancelled: int, active: int):
        """Add pie chart for lease status distribution."""
        if completed + cancelled + active == 0:
            return
            
        # Data for the chart
        chart_data = [
            ['Stav rezervácie', 'Počet'],
            ['Dokončené', completed],
            ['Zrušené', cancelled],
            ['Aktívne', active]
        ]
        
        # Write chart data starting at column E
        start_row = 4
        for i, row_data in enumerate(chart_data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=5 + j, value=value)
        
        # Create pie chart
        chart = PieChart()
        labels = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
        data = Reference(ws, min_col=6, min_row=start_row, max_row=start_row + len(chart_data) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Rozdelenie rezervácií podľa stavu"
        chart.width = 12
        chart.height = 8
        
        # Style the chart
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        ws.add_chart(chart, "E6")
    
    def _add_lease_type_chart(self, ws, private: int, business: int):
        """Add pie chart for lease type distribution."""
        if private + business == 0:
            return
            
        # Data for the chart
        chart_data = [
            ['Typ rezervácie', 'Počet'],
            ['Súkromné', private],
            ['Firemné', business]
        ]
        
        # Write chart data starting at column E, row 20
        start_row = 20
        for i, row_data in enumerate(chart_data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=5 + j, value=value)
        
        # Create pie chart
        chart = PieChart()
        labels = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
        data = Reference(ws, min_col=6, min_row=start_row, max_row=start_row + len(chart_data) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Rozdelenie rezervácií podľa typu"
        chart.width = 12
        chart.height = 8
        
        # Style the chart
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        ws.add_chart(chart, "E22")
    
    def _add_damage_chart(self, ws, damaged: int, dirty: int, exterior: int, interior: int, collisions: int):
        """Add bar chart for damage statistics."""
        # Data for the chart
        chart_data = [
            ['Typ poškodenia', 'Počet'],
            ['Vyžaduje kontrolu', damaged],
            ['Znečistené', dirty],
            ['Poškodený exteriér', exterior],
            ['Poškodený interiér', interior],
            ['Kolízie', collisions]
        ]
        
        # Write chart data starting at column H
        start_row = 4
        for i, row_data in enumerate(chart_data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=8 + j, value=value)
        
        # Create bar chart
        chart = BarChart()
        labels = Reference(ws, min_col=8, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
        data = Reference(ws, min_col=9, min_row=start_row, max_row=start_row + len(chart_data) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Štatistiky poškodení vozidiel"
        chart.width = 15
        chart.height = 10
        
        # Style the chart
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        ws.add_chart(chart, "H6")
    
    def _add_summary_insights(self, ws, lease_data: List[Dict], start_row: int):
        """Add summary insights and recommendations."""
        ws.merge_cells(f'B{start_row}:I{start_row}')
        insight_cell = ws[f'B{start_row}']
        insight_cell.value = "💡 Kľúčové poznatky a odporúčania"
        insight_cell.font = self.section_header_font
        insight_cell.fill = self.section_fill
        insight_cell.alignment = self.center_alignment
        insight_cell.border = self.medium_border
        ws.row_dimensions[start_row].height = 30
        
        insights = []
        
        # Calculate insights
        total_leases = len(lease_data)
        if total_leases > 0:
            success_rate = len([l for l in lease_data if not l['is_cancelled'] and l['time_of_return']]) / total_leases * 100
            damage_rate = len([l for l in lease_data if l['car_health_check'] or l['dirty'] or l['exterior_damage'] or l['interior_damage'] or l['collision']]) / total_leases * 100
            late_rate = len([l for l in lease_data if l['late_return']]) / total_leases * 100
            
            insights.append(f"📈 Úspešnosť dokončenia rezervácií: {success_rate:.1f}%")
            
            if damage_rate > 20:
                insights.append(f"⚠️ Vysoká miera poškodení vozidiel: {damage_rate:.1f}% - odporúčame posilniť kontroly")
            elif damage_rate < 10:
                insights.append(f"✅ Nízka miera poškodení vozidiel: {damage_rate:.1f}% - výborná starostlivosť")
            else:
                insights.append(f"📊 Miera poškodení vozidiel: {damage_rate:.1f}% - v normále")
            
            if late_rate > 15:
                insights.append(f"🚨 Vysoká miera neskorých vrátení: {late_rate:.1f}% - zvážte prísnejšie podmienky")
            elif late_rate < 5:
                insights.append(f"⭐ Nízka miera neskorých vrátení: {late_rate:.1f}% - výborná disciplína")
            else:
                insights.append(f"⏰ Miera neskorých vrátení: {late_rate:.1f}%")
        
        # Write insights
        row = start_row + 1
        for insight in insights:
            ws[f"B{row}"] = insight
            ws[f"B{row}"].font = Font(name='Calibri', size=12, color="2F5597")
            ws[f"B{row}"].fill = PatternFill("solid", "F0F8FF")
            ws[f"B{row}"].border = self.medium_border
            ws[f"B{row}"].alignment = self.left_alignment
            ws.merge_cells(f'B{row}:I{row}')
            row += 1

    def _create_leases_sheet(self, wb: Workbook, lease_data: List[Dict], year: int, month: int):
        """Create enhanced main leases data sheet with professional formatting."""
        ws = wb.create_sheet("Rezervácie")
        
        # Headers matching original format exactly
        headers = [
            "", "", "Čas od", "Čas do", "Auto", "SPZ", "Typ", "Email", 
            "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie", "Zašpinené", 
            "Pš.Interiér", "Pš.Exterier", "Kolízia"
        ]
        
        # Enhanced styling for data sheet
        # Add sheet title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f"🚗 Detailný prehľad rezervácií - {year}.{month:02d}"
        title_cell.font = Font(name='Calibri', size=20, bold=True, color="1F4E79")
        title_cell.alignment = self.center_alignment
        title_cell.fill = PatternFill("solid", "E7F3FF")
        title_cell.border = self.thick_border
        ws.row_dimensions[1].height = 35
        
        # Add one empty row
        ws.append([""] * len(headers))
        
        # Write headers in row 3 with enhanced styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            if col >= 3:  # Skip first two empty columns
                cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
                cell.alignment = self.center_alignment
                cell.fill = PatternFill("solid", "4472C4")
                cell.border = self.medium_border
        
        # Red flag cell in B3 with enhanced styling
        red_flag_cell = ws.cell(row=3, column=2)
        red_flag_cell.font = Font(bold=True, color="FFFFFF")
        red_flag_cell.fill = PatternFill("solid", "DC3545")
        red_flag_cell.border = self.medium_border
        
        # Set row height and column widths for better readability
        ws.row_dimensions[3].height = 35
        column_widths = [3, 3, 18, 18, 20, 12, 25, 25, 18, 12, 30, 12, 12, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        # Write data starting from row 4 with enhanced formatting
        row = 4
        for i, lease in enumerate(lease_data):
            # Convert dates to Slovakia timezone display format
            start_time = self.convert_to_bratislava_timezone(lease['start_of_lease'])
            end_time = self.convert_to_bratislava_timezone(lease['end_of_lease'])
            
            # Handle return time with enhanced status display
            if lease['is_cancelled']:
                return_time = "🚫 ZRUŠENÉ"
            elif lease['time_of_return']:
                return_time = self.convert_to_bratislava_timezone(lease['time_of_return'])
            else:
                return_time = "🔄 AKTÍVNE"
            
            # Prepare drive type info matching original format
            drive_info = f"{lease['gas']}, {lease['drive_type']}" if lease['gas'] and lease['drive_type'] else (lease['gas'] or lease['drive_type'] or "")
            
            # Enhanced status indicators
            def get_yes_no_indicator(value):
                return "✅ ÁNO" if value else "❌ NIE"
            
            # Create row data with enhanced indicators
            row_data = [
                "", "",  # Empty columns like original
                start_time,
                end_time,
                lease['car_name'],
                lease['stk'],
                drive_info,
                lease['email'],
                return_time,
                "⚠️ ÁNO" if lease['late_return'] else "✅ NIE",
                lease['note'],
                get_yes_no_indicator(lease['car_health_check']),
                get_yes_no_indicator(lease['dirty']),
                get_yes_no_indicator(lease['interior_damage']),
                get_yes_no_indicator(lease['exterior_damage']),
                get_yes_no_indicator(lease['collision'])
            ]
            
            # Write row with enhanced styling
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col >= 3:  # Style data columns
                    cell.font = Font(name='Calibri', size=11)
                    cell.alignment = self.center_alignment if col > 8 else self.left_alignment
                    cell.border = Border(
                        left=Side(border_style="thin", color='D0D0D0'),
                        right=Side(border_style="thin", color='D0D0D0'),
                        top=Side(border_style="thin", color='D0D0D0'),
                        bottom=Side(border_style="thin", color='D0D0D0')
                    )
            
            # Row coloring and highlighting
            if lease['is_cancelled']:
                # Cancelled leases - light red
                fill = PatternFill("solid", "FFEBEE")
            elif lease['late_return']:
                # Late returns - orange
                fill = PatternFill("solid", "FFF3E0")
            elif i % 2 == 0:
                # Alternating rows for better readability
                fill = PatternFill("solid", "F8F9FA")
            else:
                fill = PatternFill("solid", "FFFFFF")
            
            # Apply row styling
            for col in range(3, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                
                # Special highlighting for problem indicators
                if col >= 12:  # Damage columns
                    if "✅ ÁNO" in str(cell.value):
                        cell.fill = PatternFill("solid", "FFCDD2")
                        cell.font = Font(name='Calibri', size=11, bold=True, color="B71C1C")
            
            ws.row_dimensions[row].height = 25
            row += 1
        
        # Add summary row at the bottom
        if lease_data:
            row += 1
            ws.merge_cells(f'A{row}:P{row}')
            summary_cell = ws[f'A{row}']
            summary_cell.value = f"📊 Celkom: {len(lease_data)} rezervácií | ✅ Dokončené: {len([l for l in lease_data if not l['is_cancelled'] and l['time_of_return']])} | ❌ Zrušené: {len([l for l in lease_data if l['is_cancelled']])} | 🔄 Aktívne: {len([l for l in lease_data if not l['is_cancelled'] and not l['time_of_return']])}"
            summary_cell.font = Font(name='Calibri', size=12, bold=True, color="1F4E79")
            summary_cell.alignment = self.center_alignment
            summary_cell.fill = PatternFill("solid", "E3F2FD")
            summary_cell.border = self.thick_border
            ws.row_dimensions[row].height = 30

    def generate_previous_month_report(self, force_rebuild: bool = False) -> str:
        """
        Generate report for the previous month.
        This is the method that would typically be called by cron jobs.
        """
        current_date = self.get_sk_date()
        
        # Calculate previous month
        if current_date.month == 1:
            prev_month = 12
            prev_year = current_date.year - 1
        else:
            prev_month = current_date.month - 1
            prev_year = current_date.year
        
        print(f"INFO: Generating report for {prev_year}-{prev_month:02d}")
        return self.create_excel_report(prev_year, prev_month, force_rebuild)
    
    def generate_current_month_report(self, force_rebuild: bool = False) -> str:
        """
        Generate report for the current month.
        Useful for testing or mid-month reports.
        """
        current_date = self.get_sk_date()
        print(f"INFO: Generating current month report for {current_date.year}-{current_date.month:02d}")
        return self.create_excel_report(current_date.year, current_date.month, force_rebuild)
    
    def run_monthly_task(self):
        """
        Run the complete monthly task: generate report and clean old ones.
        This is the main method for automatic Docker execution.
        """
        try:
            print(f"INFO: Starting monthly Excel report generation at {self.get_sk_date()}")
            
            # Generate previous month report
            filename = self.generate_previous_month_report()
            print(f"SUCCESS: Generated report: {filename}")
                        
            print("INFO: Monthly task completed successfully")
            
        except Exception as e:
            print(f"ERROR: Monthly task failed: {e}")
            raise


def main():
    """
    Main function for automatic monthly report generation.
    Runs without arguments for Docker container automation.
    """
    writer = MonthlyExcelWriter()
    
    try:
        # Run the complete monthly task automatically
        writer.run_monthly_task()
            
    except Exception as e:
        print(f"ERROR: Failed to generate monthly report: {e}")
        exit(1)


if __name__ == "__main__":
    main()