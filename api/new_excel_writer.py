# This new version should only be triggered once a month, at 00:05 the day of the next month
# It will still generate the same excel, but now the leases will be ordered by when they actually began
# Also it will have a front page with statistics, most used car, most leases done, most damage by user etc.
# This will be called by a cron job, so no while loop hogging the thread
#* The cron job can be set in a docker linux container


# The files will now be called just by their year and month: 2025.04 ICLS Report.xlsx
# No checks will be made if another file exists as the cron job will take care of that

# ? Maybe implement a force rebuild function, so the admin can generate a report whenver, if the cron job fucks itself


import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook
from datetime import datetime
import os
import psycopg2
import pytz


db_host = "127.0.0.1"
db_port = 5434
db_user = "postgres"
db_pass = "admin123"
db_name = "postgres"

class ExcelWriter:
    # Define styles
    red_flag_ft = Font(bold=True, color="B22222")
    red_flag_fill = PatternFill("solid", "B22222")
    Header_fill = PatternFill("solid", "00CCFFFF")
    Header_ft = Font(bold=True, color="000000", size=20)
    Data_ft = Font(size=17)  # New font for data cells

    Header_border = Border(
        left=Side(border_style="medium", color='FF000000'),
        right=Side(border_style="medium", color='FF000000'),
        top=Side(border_style="medium", color='FF000000'),
        bottom=Side(border_style="medium", color='FF000000')
    )

    header_alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    def connect_to_db(self):
        try:
            db_con = psycopg2.connect(dbname=db_name, user=db_user, host=db_host, port=db_port, password=db_pass)
            cur = db_con.cursor()
            return db_con, cur
        except psycopg2.Error as e:
            print(f"cannot login to db {e}")
            return None, str(e)
    
    def get_month_lengths(self, month,leap_year=False) -> int:
        lengths = {
            1: 31,
            2: 29 if leap_year else 28,
            3: 31,
            4: 30,
            5: 31,
            6: 30,
            7: 31,
            8: 31,
            9: 30,
            10: 31,
            11: 30,
            12: 31
        }
        return lengths[month]

    def isLeapYear(self, year:int) ->bool:
        if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
            return True
        else:
            return False

    def convert_to_datetime(self, string) -> datetime:
        try:
            # Parse string, handling timezone if present
            dt_obj = datetime.strptime(string, "%Y-%m-%d %H:%M:%S")
            return dt_obj
        except Exception as e: 
            raise ValueError(f"Invalid datetime format: {string}") from e

    def get_sk_date(self) -> datetime:
        bratislava_tz = pytz.timezone('Europe/Bratislava')

        dt_obj = datetime.now()
        utc_time = dt_obj.replace(tzinfo=pytz.utc) if dt_obj.tzinfo is None else dt_obj.astimezone(pytz.utc)
        bratislava_time = utc_time.astimezone(bratislava_tz)  # Convert to Bratislava timezone
        return bratislava_time
    
    # Get lease data from the DB, take only from the last month
    # Order it by their starting day, if a value does not exist, replace with NULL
    #* You need to generate the last month date yourself, as sql is retatrded
    def get_lease_data(self) -> list:
        conn, curr = self.connect_to_db()

        today = self.get_sk_date()

        #! TEMPORARY remove later
        today = datetime(
            year= 2025,
            month= 5,
            day=1,
            hour=0,
            minute=5
        )
        # Construct the last month manually, if today is a new year, the last month has to be one year shorter
        # This will act as the first interval, while the second begins the next month at 00:05
        last_month = datetime(
            year= today.year      if (today.month != 1) else today.year-1,
            month= today.month -1 if (today.month != 1) else 12,
            day=1,
            hour=0,
            minute=0
        )

        try:
            query = """
                SELECT 
                    l.start_of_lease, l.end_of_lease, l.time_of_return, l.note, l.private, l.car_health_check, l.dirty, l.exterior_damage, l.interior_damage, l.collision,
                    c.name, c.stk, c.drive_type, c.gas,
                    d.email       
                FROM lease AS l
                JOIN car AS c 
                    ON l.id_car = c.id_car
                JOIN driver AS d 
                    ON l.id_driver = d.id_driver
                WHERE start_of_lease > %s AND start_of_lease < %s
                ORDER BY l.start_of_lease ASC;

                """
            curr.execute(query, (last_month, today, ))
            res = curr.fetchall()
        except Exception as e:
            raise RuntimeError(f"{e}, During SQL operation")
        
        conn.close()
        return res




    # Create the excel file
    def createExcel(self) -> object:
        wb = Workbook(write_only=False)
        del wb["Sheet"]
        current_time = self.get_sk_date()


        current_month = self.get_month_lengths( 
                int(current_time.month), 
                self.isLeapYear(current_time.year) 
            )

        for day in range(1, current_month):

            ws = wb.create_sheet(str(day))
            filler = ["","","","","","","",""]
            data = [filler,filler,["", "", "Čas od", "Čas do", "Auto", "SPZ", "Typ","Email", "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie","Zašpinené", "Pš.Interiér", "Pš.Exterier", "Kolizia"]]

            for row in data:
                ws.append(row)

            # Format red flag cell (B3)
            red_flag_cell = ws["B3"]
            red_flag_cell.font   = self.red_flag_ft
            red_flag_cell.fill   = self.red_flag_fill
            red_flag_cell.border = self.Header_border
            email_cell           = ws["B3"]


            # Set row height for header row
            ws.row_dimensions[3].height = 35

            # Set column widths for data columns
            for col in ["C", "D", "E", "F", "G", "H", "I", "J","K", "L", "M", "N", "O", "P"]:
                ws.column_dimensions[col].width = 23

            # Format header row (C3:J3)
            for row_cells in ws["C3:P3"]:
                for cell in row_cells:
                    cell.font      = self.Header_ft
                    cell.alignment = self.header_alignment
                    cell.fill      = self.Header_fill
                    cell.border    = self.Header_border


        # List: (datetime.datetime(2025, 2, 14, 13, 30, tzinfo=datetime.timezone.utc), datetime.datetime(2025, 2, 17, 8, 30, tzinfo=datetime.timezone.utc), None, None, False, None, False, False, False, False, 'Wolksvagen Golf', 'BB676GF', 'AUTOMAT', 'BENZ??N', 'test@manager.sk')
        #l.start_of_lease, l.end_of_lease, l.time_of_return, l.note, l.private, l.car_health_check, l.dirty, l.exterior_damage, l.interior_damage, l.collision, c.name, c.stk, c.drive_type, c.gas, d.email
        # "", "", "Čas od", "Čas do", "Auto", "SPZ", "Typ","Email", "Odovzdanie", "Meškanie", "Poznámka", "Poškodenie","Zašpinené", "Pš.Interiér", "Pš.Exterier", "Kolizia"
        data = self.get_lease_data()
        
        #! This will be a problem, if we have a multi month lease, where should it be 
        for ws_title in range(1, current_month):
            
            ws_title = str(ws_title)
            if ws_title not in wb.sheetnames:
                pass
            else:
                worksheet = wb[ws_title]
                for row in data:
         
                    if row[0].day != int(ws_title):
                        pass
                    else:
                        start_of_lease = row[0].strftime("%d/%m/%Y  %H:%M")
                        end_of_lease   = row[0].strftime("%d/%m/%Y  %H:%M")
                        time_of_return = row[0].strftime("%d/%m/%Y  %H:%M:%S")
                        
                        exc_row = ["","",
                                   start_of_lease, 
                                   end_of_lease, 
                                   row[10], 
                                   row[11],
                                   f"{row[12]}, {row[13]}", 
                                   row[14], 
                                   "3","NULL","NULL","NULL","NULL","NULL","NULL","NULL"
                                   ]
                        worksheet.append(exc_row)

            
        for worksheet in wb.worksheets:
            for row in worksheet.iter_rows(min_row=3):
                status_cell = row[8]  # Column I (0-based index 8) for "Meškanie"
                if str(status_cell.value) == "3":
                    # Check the first two cells (columns A and B) in the row
                    for cell in row[0:2]:  # Indices 0 (A) and 1 (B)
                        if cell.value is None or cell.value == "":
                            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            cell.fill = yellow_fill
                            break  # Color the first empty cell and exit

        wb_name = f"{current_time.year}.{current_time.month} ICLS Report.xlsx"
        wb.save(f"{os.getcwd()}/reports{wb_name}")


obj = ExcelWriter()

obj.createExcel()